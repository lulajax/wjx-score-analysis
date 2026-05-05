"""新学员登记：从消息文本中识别学员字段，追加到 Excel 登记表"""

import json
import os
import re
from datetime import datetime

import openpyxl

# --- 默认字段配置（field_config.json 不存在时使用） ---
_DEFAULT_CONFIG = [
    {"key": "协议是否签署", "excel_header": "协议是否签署", "aliases": ["协议", "签署协议", "是否签署协议", "合同"], "is_date": False},
    {"key": "学员姓名", "excel_header": "学员姓名", "aliases": ["姓名", "名字", "名称"], "is_date": False},
    {"key": "性别", "excel_header": "性别", "aliases": [], "is_date": False},
    {"key": "年龄", "excel_header": "年龄", "aliases": [], "is_date": False},
    {"key": "专业", "excel_header": "专业", "aliases": [], "is_date": False},
    {"key": "学历和学位", "excel_header": "学历和学位", "aliases": ["学历", "学位", "学历学位", "最高学历"], "is_date": False},
    {"key": "学校", "excel_header": "学校", "aliases": ["毕业院校", "院校", "大学"], "is_date": False},
    {"key": "从事何种行业", "excel_header": "从事何种行业", "aliases": ["行业", "职业", "工作", "从事行业", "目前职业", "从事工作"], "is_date": False},
    {"key": "备考地区", "excel_header": "备考地区", "aliases": ["考试地区", "地区", "所考省份", "报考省份"], "is_date": False},
    {"key": "目标考试", "excel_header": "目标考试", "aliases": ["考试", "目标"], "is_date": False},
    {"key": "报考年份", "excel_header": "报考年份", "aliases": [], "is_date": False},
    {"key": "参考经历", "excel_header": "参考经历（0基础/备考几月）", "aliases": ["参考经历（0基础/备考几月）", "参考经历(0基础/备考几月)", "参考经历（0基础）", "考试经历", "备考经历", "基础"], "is_date": False},
    {"key": "每日学习时长", "excel_header": "每日学习时长", "aliases": ["学习时长", "学习时间", "每天学习时长", "每天学习时间"], "is_date": False},
    {"key": "政治面貌", "excel_header": "政治面貌", "aliases": [], "is_date": False},
    {"key": "毕业时间", "excel_header": "毕业时间", "aliases": ["毕业日期"], "is_date": True},
    {"key": "邮寄地址", "excel_header": "邮寄地址", "aliases": ["地址", "收件地址", "快递地址", "收货地址"], "is_date": False},
    {"key": "电话", "excel_header": "电话", "aliases": ["手机", "手机号", "手机号码", "联系电话", "联系方式", "电话号码", "tel"], "is_date": False},
    {"key": "报名课程产品名称", "excel_header": "报名课程产品名称（全名称）", "aliases": ["报名课程产品名称（全名称）", "报名课程产品名称(全名称)", "课程名称", "课程产品", "课程", "报名课程", "产品名称"], "is_date": False},
    {"key": "班型", "excel_header": "班型", "aliases": [], "is_date": False},
    {"key": "实付金额", "excel_header": "实付金额", "aliases": ["金额", "付款金额", "缴费金额", "费用", "学费"], "is_date": False},
    {"key": "业绩归属", "excel_header": "业绩归属", "aliases": ["业绩", "归属"], "is_date": False},
    {"key": "报班时间", "excel_header": "报班时间", "aliases": ["报班日期", "报名时间", "报名日期", "入学时间"], "is_date": True},
    {"key": "学员特殊情况备注", "excel_header": "备注", "aliases": ["备注", "特殊情况备注", "特殊情况", "特殊备注"], "is_date": False},
]

# --- 运行时状态 ---
_config = list(_DEFAULT_CONFIG)
_config_path = None

# 行首编号/标点前缀：1. 或 18: 或 18： 或 ① 或 - 或 · 等
_LINE_PREFIX_RE = re.compile(r"^\s*(?:\d{1,2}\s*[.．、:：)\]）】]\s*|[①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲⑳]\s*|[-·•]\s+)")

# 噪音行模式
_NOISE_RE = re.compile(
    r"^(老师|你好|您好|嗨|hi|hello|麻烦|请|谢谢|感谢|辛苦|好的|收到|以下是|信息如下|学员信息|新学员|报名信息)",
    re.IGNORECASE,
)

# 值的模式识别正则
_PHONE_RE = re.compile(r"^1[3-9]\d{9}$")
_MONEY_RE = re.compile(r"^\d{3,6}(?:\.\d{1,2})?(?:元)?$")
_GENDER_RE = re.compile(r"^[男女]$")
_AGE_RE = re.compile(r"^\d{1,2}岁?$")

# 运行时查找结构（由 _reload_runtime() 填充）
FIELD_MAP = {}
_FIELD_ALIASES = {}
DATE_FIELDS = set()
FIELD_ORDER = []
_SORTED_ALIASES = []


def _build_runtime(cfg):
    """从配置列表构建运行时查找结构"""
    field_map = {}
    aliases = {}
    date_fields = set()
    field_order = []
    for item in cfg:
        key = item["key"]
        excel_header = item.get("excel_header", key)
        field_map[key] = excel_header
        field_order.append(key)
        aliases[key] = key  # key 本身也是别名
        for alias in item.get("aliases", []):
            if alias:
                aliases[alias] = key
        if item.get("is_date", False):
            date_fields.add(key)
    return field_map, aliases, date_fields, field_order


def _reload_runtime():
    global FIELD_MAP, _FIELD_ALIASES, DATE_FIELDS, FIELD_ORDER, _SORTED_ALIASES
    FIELD_MAP, _FIELD_ALIASES, DATE_FIELDS, FIELD_ORDER = _build_runtime(_config)
    _SORTED_ALIASES = sorted(_FIELD_ALIASES.keys(), key=len, reverse=True)


# 初始构建（使用默认配置）
_reload_runtime()


def load_field_config(path):
    """从 JSON 文件加载字段配置，文件不存在时写出默认配置"""
    global _config, _config_path
    _config_path = path
    if os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            _config = json.load(f)
    else:
        save_field_config(path, _DEFAULT_CONFIG)
        _config = list(_DEFAULT_CONFIG)
    _reload_runtime()


def save_field_config(path, cfg):
    """将字段配置写入 JSON 文件，并更新运行时状态"""
    global _config, _config_path
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)
    _config = list(cfg)
    _config_path = path
    _reload_runtime()


def get_field_config():
    """返回当前字段配置列表（副本）"""
    return list(_config)


def _parse_one(lines):
    """从一组行中解析一个学员的字段"""
    result = {}
    unmatched = []

    for line in lines:
        line = line.strip()
        if not line:
            continue

        plain = _LINE_PREFIX_RE.sub("", line)
        if _NOISE_RE.match(plain):
            continue

        line = plain
        matched_field = None
        rest = line

        for alias in _SORTED_ALIASES:
            if line.startswith(alias):
                matched_field = _FIELD_ALIASES[alias]
                rest = line[len(alias):]
                break

        if matched_field:
            rest = re.sub(r"^[\s:：=\-]+", "", rest).strip()
            result[matched_field] = rest
        else:
            unmatched.append(line)

    for line in unmatched:
        val = re.sub(r"^[\s:：=\-]+", "", line).strip()
        if not val:
            continue
        if "电话" not in result and _PHONE_RE.match(val):
            result["电话"] = val
        elif "性别" not in result and _GENDER_RE.match(val):
            result["性别"] = val
        elif "年龄" not in result and _AGE_RE.match(val):
            result["年龄"] = val
        elif "实付金额" not in result and _MONEY_RE.match(val.replace(",", "").replace("元", "")):
            result["实付金额"] = val

    return result


def parse_message(text):
    """从消息文本中解析学员字段。支持多人批量粘贴，返回列表 [{字段名: 值}, ...]"""
    lines = text.strip().splitlines()

    name_positions = []
    for i, line in enumerate(lines):
        stripped = _LINE_PREFIX_RE.sub("", line.strip())
        for alias in ("学员姓名", "姓名"):
            if stripped.startswith(alias):
                name_positions.append(i)
                break

    if len(name_positions) <= 1:
        result = _parse_one(lines)
        return [result] if result else []

    groups = []
    for idx, pos in enumerate(name_positions):
        end = name_positions[idx + 1] if idx + 1 < len(name_positions) else len(lines)
        start = name_positions[idx - 1] if idx > 0 else 0
        if idx > 0:
            block_start = pos
            for j in range(name_positions[idx - 1] + 1, pos):
                stripped = lines[j].strip()
                if not stripped:
                    block_start = j + 1
                    break
                m = re.match(r"^\s*1\s*[.．、:：]", stripped)
                if m:
                    block_start = j
                    break
            start = block_start
        groups.append(lines[start:end])

    results = []
    for group in groups:
        parsed = _parse_one(group)
        if parsed:
            results.append(parsed)
    return results


def _parse_date(text):
    text = text.strip()
    for fmt in ("%Y年%m月%d日", "%Y-%m-%d", "%Y年%m月", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return text


def _find_next_row(ws):
    for r in range(2, ws.max_row + 2):
        if ws.cell(r, 1).value is None or str(ws.cell(r, 1).value).strip() in ("", "\n"):
            return r
    return ws.max_row + 1


def _get_header_map(ws):
    hmap = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            hmap[v.strip()] = c
    return hmap


def register(xlsx_path, data):
    """将学员数据追加到 Excel 文件，返回写入结果"""
    xlsx_path = os.path.abspath(xlsx_path)

    if not os.path.exists(xlsx_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        for i, header in enumerate(FIELD_MAP.values(), 1):
            ws.cell(1, i, header)
        wb.save(xlsx_path)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    header_map = _get_header_map(ws)
    next_col = max(header_map.values()) + 1 if header_map else 1
    row = _find_next_row(ws)

    filled = []
    for user_field, value in data.items():
        if not value:
            continue
        excel_header = FIELD_MAP.get(user_field, user_field)

        if excel_header not in header_map:
            header_map[excel_header] = next_col
            ws.cell(1, next_col, excel_header)
            next_col += 1

        col = header_map[excel_header]

        if user_field in DATE_FIELDS:
            value = _parse_date(str(value))
            if isinstance(value, datetime):
                ws.cell(row, col, value).number_format = "YYYY-MM-DD"
                filled.append({"field": excel_header, "value": value.strftime("%Y-%m-%d")})
                continue

        if user_field == "实付金额":
            try:
                value = float(str(value).replace(",", "").replace("元", ""))
            except ValueError:
                pass

        ws.cell(row, col, value)
        filled.append({"field": excel_header, "value": str(value)})

    wb.save(xlsx_path)
    return {"row": row, "count": len(filled), "fields": filled, "path": xlsx_path}
