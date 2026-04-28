#!/usr/bin/env python3
"""
展鹏公考 - 新学员登记
将学员信息追加到 新学员登记表.xlsx 的下一空行。

用法:
  python3 register_student.py --xlsx 新学员登记表.xlsx --data '{...}'
"""

import json, sys, io, os, argparse
from datetime import datetime

# Windows 终端中文输出
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")

try:
    import openpyxl
except ImportError:
    print("正在安装 openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl


# 用户字段 → Excel 表头 的映射
FIELD_MAP = {
    "合同是否签署":       "合同是否签署",
    "入学时间":           "入学时间",
    "学员姓名":           "学员姓名",
    "报考年份":           "报考年份",
    "班型":               "班型（报名课程产品名称全名称）",
    "报名课程产品名称":   "班型（报名课程产品名称全名称）",
    "联系方式":           "联系方式",
    "电话":               "联系方式",
    "实付金额":           "实付金额",
    "性别":               "性别",
    "年龄":               "年龄",
    "学历":               "学历",
    "学校":               "学校",
    "毕业时间":           "毕业时间",
    "专业":               "专业",
    "所考省份":           "所考省份",
    "目标考试":           "目标考试",
    "参考经历":           "参考经历（0基础/备考几月/几年）",
    "收货地址":           "收货地址",
    "从事何种行业":       "从事何种行业（在职几年/在校）",
    "备注":               "备注",
}

# 需要按日期处理的字段
DATE_FIELDS = {"入学时间", "毕业时间"}


def parse_date(text):
    """尝试将日期文本解析为 datetime 对象"""
    text = text.strip()
    for fmt in ("%Y年%m月%d日", "%Y-%m-%d", "%Y年%m月", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return text  # 无法解析则原样写入


def find_next_row(ws):
    """找到第一个整行全空的行"""
    for r in range(2, ws.max_row + 2):
        row_empty = True
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip() not in ("", "\n"):
                row_empty = False
                break
        if row_empty:
            return r
    return ws.max_row + 1


def get_header_map(ws):
    """读取第 1 行，返回 {表头名: 列号}"""
    hmap = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            hmap[v.strip()] = c
    return hmap


def main():
    ap = argparse.ArgumentParser(description="新学员登记")
    ap.add_argument("--xlsx", required=True, help="Excel 文件路径")
    ap.add_argument("--data", required=True, help="学员信息 JSON 字符串")
    args = ap.parse_args()

    data = json.loads(args.data)
    xlsx_path = os.path.abspath(args.xlsx)
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    header_map = get_header_map(ws)
    next_col = max(header_map.values()) + 1 if header_map else 1
    row = find_next_row(ws)

    filled = []
    for user_field, value in data.items():
        excel_header = FIELD_MAP.get(user_field, user_field)

        # 如果表头不存在，新增列
        if excel_header not in header_map:
            header_map[excel_header] = next_col
            ws.cell(1, next_col, excel_header)
            print(f"  新增列 {openpyxl.utils.get_column_letter(next_col)}: {excel_header}")
            next_col += 1

        col = header_map[excel_header]

        # 日期字段特殊处理
        if user_field in DATE_FIELDS:
            value = parse_date(str(value))
            if isinstance(value, datetime):
                ws.cell(row, col, value).number_format = "YYYY-MM-DD"
                filled.append(f"  {excel_header} ({openpyxl.utils.get_column_letter(col)}): {value.strftime('%Y-%m-%d')}")
                continue

        # 实付金额转数字
        if user_field == "实付金额":
            try:
                value = float(str(value).replace(",", "").replace("元", ""))
            except ValueError:
                pass

        ws.cell(row, col, value)
        filled.append(f"  {excel_header} ({openpyxl.utils.get_column_letter(col)}): {value}")

    wb.save(xlsx_path)
    print(f"\n已写入第 {row} 行，共 {len(filled)} 个字段:")
    for f in filled:
        print(f)
    print(f"\n文件已保存: {xlsx_path}")


if __name__ == "__main__":
    main()
